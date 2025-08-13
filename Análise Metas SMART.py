import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

# Caminho do arquivo original
arquivo = r"C:\Users\cferlin\Downloads\Relatório de Avaliações 2025 (2).xlsx"

# Leitura da aba metas
df = pd.read_excel(arquivo, sheet_name="Metas")

# Verificar colunas esperadas
colunas_necessarias = {'Título', 'Descrição', 'Colaborador'}
colunas_atuais = set(df.columns)
colunas_faltando = colunas_necessarias - colunas_atuais
if colunas_faltando:
    raise ValueError(f"Faltam as colunas: {colunas_faltando}")

# SMART refinado
def is_specific(texto):
    return any(p in texto.lower() for p in ["melhorar", "promover", "agendar", "originar", "establish", "revisar", "elaborar", "reagir", "maximizar", "entregar", "conciliar", "aprimorar", "implementar", "reduzir", "aumentar", "criar", "desenvolver", "acompanhar", "identificar", "efetivar", "mitigar", "organizar", "participar", "manter", "processeguir", "apoiar", "planejar", "avaliar", "gerenciar", "alcançar", "auxiliar", "interagir", "mapear", "assegurar", "fortalecer", "confeccionar", "analisar", "garantir", "realizar", "adequar", "controlar", "otimizar", "executar"])

def is_measurable(texto):
    return any(p in texto.lower() for p in ["%", "percentual", "indicador", "índice", "meta de", "redução de", "menores ou iguais", "pelo menos", "atingir o tempo", "aumento de", "quantidade", "criar", "até", "fazer", "realizou", "realização", "apresentar", "criar", "no mínimo", "no minimo", "não atendeu as expectativas", "gerar um", "abaixo das expectativas", "atendeu as expectativas", "acima das expectativas", "excedeu as expectativas", "mensurar", "volume", "número"])

def is_achievable(texto):
    return not any(p in texto.lower() for p in ["impossível", "inviável", "utópico", "perfeito", "totalmente perfeito", "100% garantido", "sem falhas", "infalível", "atingir a perfeição", "ser referência mundial", "superar todas as expectativas", "todos os clientes satisfeitos", "atingir excelência absoluta", "ser o número 1 do mercado", "liderança total", "zerar problemas", "eliminar completamente", "controle total", "crescimento ilimitado", "resultado máximo garantido"])

def is_relevant(texto):
    return any(p in texto.lower() for p in ["cliente", "safra", "grãos", "viterra", "negócio", "CIA", "companhia", "empresa", "resultado", "indicador", "estratégia", "eficiência", "qualidade", "processo", "área"])

def is_time_bound(texto):
    return any(p in texto.lower() for p in ["até", "prazo", "mensal", "trimestral", "anual", "em x dias", "em x semanas", "em x meses", "semanal", "em 2025"])

def gerar_sugestoes(validacoes):
    sugestoes = []
    if not validacoes["Específica"]:
        sugestoes.append("Use um verbo claro e objetivo.")
    if not validacoes["Mensurável"]:
        sugestoes.append("Adicione uma métrica ou indicador.")
    if not validacoes["Atingível"]:
        sugestoes.append("Torne a meta mais realista.")
    if not validacoes["Relevante"]:
        sugestoes.append("Conecte a meta à área ou negócio.")
    if not validacoes["Temporal"]:
        sugestoes.append("Estabeleça um prazo ou marco temporal.")
    return " | ".join(sugestoes)

# Processamento
fora_do_padrao = []
dentro_do_padrao = []

for _, row in df.iterrows():
    titulo = str(row['Título']) if pd.notna(row['Título']) else ""
    descricao = str(row['Descrição']) if pd.notna(row['Descrição']) else ""
    texto = f"{titulo} {descricao}"

    validacoes = {
        "Específica": is_specific(texto),
        "Mensurável": is_measurable(texto),
        "Atingível": is_achievable(texto),
        "Relevante": is_relevant(texto),
        "Temporal": is_time_bound(texto)
    }

    registro = {
        "Colaborador": row['Colaborador'],
        "Título da Meta": titulo,
        "Descrição da Meta": descricao,
        **validacoes
    }

    if all(validacoes.values()):
        dentro_do_padrao.append(registro)
    else:
        registro["Sugestão de Melhoria"] = gerar_sugestoes(validacoes)
        fora_do_padrao.append(registro)

# Exportar os dados
arquivo_fora = "metas_fora_do_padrao_smart_com_sugestoes.xlsx"
arquivo_dentro = "metas_adequadas_ao_padrao_smart.xlsx"

df_fora = pd.DataFrame(fora_do_padrao)
df_dentro = pd.DataFrame(dentro_do_padrao)

df_fora.to_excel(arquivo_fora, index=False)
df_dentro.to_excel(arquivo_dentro, index=False)

# Formatar condicionalmente o arquivo com metas fora do padrão
wb = load_workbook(arquivo_fora)
ws = wb.active

# Estilos de preenchimento
vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Aplica nas colunas dos critérios SMART (assumindo ordem fixa após as 3 primeiras colunas)
colunas_smart = ["Específica", "Mensurável", "Atingível", "Relevante", "Temporal"]
col_idx = {cell.value: idx+1 for idx, cell in enumerate(ws[1]) if cell.value in colunas_smart}

for col in colunas_smart:
    idx = col_idx[col]
    letra = chr(64 + idx)
    ws.conditional_formatting.add(
        f"{letra}2:{letra}{ws.max_row}",
        CellIsRule(operator='equal', formula=['FALSE'], fill=vermelho)
    )
    ws.conditional_formatting.add(
        f"{letra}2:{letra}{ws.max_row}",
        CellIsRule(operator='equal', formula=['TRUE'], fill=verde)
    )

wb.save(arquivo_fora)

print("✅ Arquivos gerados com formatação:")
print(f"- {arquivo_fora} (com destaque SMART)")
print(f"- {arquivo_dentro}")
