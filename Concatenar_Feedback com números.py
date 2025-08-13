import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Carregar o arquivo Excel
df = pd.read_excel(r'C:\Users\cferlin\Downloads\Feedbacks_fechamento24.xlsx')

# Verificar se as colunas necessárias estão no DataFrame
if "Destinatário" in df.columns and "Comentário" in df.columns:
    # Agrupar pelos valores duplicados na coluna "Destinatário"
    df_concatenado = df.groupby("Destinatário", as_index=False).agg({
        "Comentário": lambda x: " || ".join(f"**{i + 1}** {coment}" for i, coment in enumerate(filter(pd.notnull, x)))
    })

    # Salvar o resultado em um novo arquivo Excel
    arquivo_saida = "resultado_concatenado_feedback24.xlsx"
    df_concatenado.to_excel(arquivo_saida, index=False, engine="openpyxl")

    # Abrir o arquivo Excel para aplicar a formatação em negrito
    wb = load_workbook(arquivo_saida)
    ws = wb.active

    # Aplicar o negrito aos números na coluna de comentários
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            formatted_text = cell.value.split(" | ")
            formatted_parts = []
            for part in formatted_text:
                if "**" in part:
                    num, text = part.split(" ", 1)
                    formatted_parts.append((num.replace("**", ""), text))
        wb.save(arquivo_saida)
    print(f"Arquivo salvo com negrito como {arquivo_saida}")
else:
    print("As colunas 'Destinatário' e 'Comentário' não foram encontradas no arquivo Excel.")
