import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Carregar o arquivo Excel
df = pd.read_excel(r'C:\Users\cferlin\Downloads\Base_Comentários.xlsx')

# Verificar se as colunas necessárias estão no DataFrame
required_columns = [
    "Avaliação de Desempenho:Colaborador",
    "1ª Autoavaliação do Colaborador",
    "1ª Avaliação do Gestor",
    "2ª Autoavaliação do Colaborador",
    "2ª Avaliação do Gestor",
    "Motivo do Cancelamento"
]

if all(col in df.columns for col in required_columns):
    # Agrupar pelos valores duplicados na coluna "Avaliação de Desempenho:Colaborador"
    def concatenate_comments(series):
        return " || ".join(f"**{i + 1}** {coment}" for i, coment in enumerate(series.dropna()))


    df_concatenado = df.groupby("Avaliação de Desempenho:Colaborador", as_index=False).agg({
        "1ª Autoavaliação do Colaborador": concatenate_comments,
        "1ª Avaliação do Gestor": concatenate_comments,
        "2ª Autoavaliação do Colaborador": concatenate_comments,
        "2ª Avaliação do Gestor": concatenate_comments,
        "Motivo do Cancelamento": concatenate_comments
    })

    # Salvar o resultado em um novo arquivo Excel
    arquivo_saida = "resultado_concatenado_comentários.xlsx"
    df_concatenado.to_excel(arquivo_saida, index=False, engine="openpyxl")

    # Abrir o arquivo Excel para aplicar a formatação em negrito
    wb = load_workbook(arquivo_saida)
    ws = wb.active

    # Aplicar o negrito aos números formatados em "**n**"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            if cell.value:
                parts = cell.value.split(" || ")
                formatted_parts = []
                for part in parts:
                    if part.startswith("**") and "**" in part:
                        num, text = part.split("** ", 1)
                        formatted_parts.append((num.strip("**"), text.strip()))
                if formatted_parts:
                    cell.font = Font(bold=True)

    # Salvar o arquivo Excel atualizado
    wb.save(arquivo_saida)
    print(f"Arquivo salvo com negrito como {arquivo_saida}")
else:
    print("As colunas necessárias não foram encontradas no arquivo Excel.")
